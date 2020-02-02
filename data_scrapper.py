import requests
from bs4 import BeautifulSoup
import bs4
import itertools
import pandas as pd
import os
from os import path
from datetime import datetime
from openpyxl import load_workbook
import tqdm
import time
import collections
import sys

class league_link_standings():
    def __init__(self):
        #STANDINGS - Get dictionary with the leagues and the links for those leagues for the standings (www.soccerstats.com)  
        league_link_dict_std = {}
        search_soup = BeautifulSoup(requests.get('https://www.soccerstats.com/', headers={'User-Agent': 'Mozilla/5.0'}).content, 'lxml').find('table',{'cellspacing':'1'}).find_all('span')
        print('\nGetting links for Standings')
        for i, league in zip(tqdm.tqdm(range(len(search_soup))), search_soup):
            league_link_dict_std[league['title']] = league.find('a')['href']
            time.sleep(0.1)
        self.dict_standings = league_link_dict_std
    
    def get_links_standings(self):
        return self.dict_standings

class league_link_players():
    def __init__(self):
        #SCORERS, ASSISTANTS AND GOALKEEPERS - www.transfermarkt.com
        
        # From this line on this class, all the lines marked with the hash mean to be the objects and processes by which the program gets the results specified below the lines '###RESULT OF ABOVE###' 
        # that you can see at the end of every paragraph. I decided to dispense with all this processes just for saving some computer resources, as getting all the necessary elements to build the 
        # Transfermarkt links could be a long and tedious process for the CPU; but to keep all the lines in the document for demonstrating the way how I get the last two elements: 'league_name_code_list'
        # and 'league_name_code_season'    
        
        #countries_link_dict = {}
        #countries_dict = {}
        league_link_dict_std = {}
        search_soup = BeautifulSoup(requests.get('https://www.soccerstats.com/', headers={'User-Agent': 'Mozilla/5.0'}).content, 'lxml').find('table',{'cellspacing':'1'}).find_all('span')
        for league in search_soup:
            league_link_dict_std[league['title']] = league.find('a')['href']
            #countries_link_dict[league['title'].partition(' ')[0]] = ''
            #country = league['title'].partition(' ')[0]
            #if country not in countries_dict.keys():
             #   countries_dict[country] = 0
            #if country in countries_dict.keys():
             #   countries_dict[country] = countries_dict[country]+1
        ###RESULT OF ABOVE###
        #countries_link_dict = {'Argentina': '', 'Austria': '', 'Australia': '', 'Belgium': '', 'Brazil': '', 'Switzerland': '', 'Germany': '', 'Denmark': '', 'England': '', 'Spain': '', 'Finland': '', 'France': '', 'Netherlands': '', 'Italy': '', 'Japan': '', 'Norway': '', 'Poland': '', 'Portugal': '', 'Russia': '', 'Scotland': '', 'Sweden': '', 'Turkey': ''}
        #countries_dict = {'Argentina': 1, 'Austria': 1, 'Australia': 1, 'Belgium': 1, 'Brazil': 1, 'Switzerland': 1, 'Germany': 2, 'Denmark': 1, 'England': 5, 'Spain': 2, 'Finland': 1, 'France': 2, 'Netherlands': 1, 'Italy': 2, 'Japan': 1, 'Norway': 1, 'Poland': 1, 'Portugal': 1, 'Russia': 1, 'Scotland': 1, 'Sweden': 1, 'Turkey': 1}
        
        #Get league_name, league_code and year
        print('\nGetting links for Scorers, Assistants and Goalkeepers')
        #for country, num_leagues in countries_dict.items():
         #   countries_dict[country] = list(range(0,num_leagues))
        ###RESULT OF ABOVE###
        #countries_dict = {'Argentina': [0], 'Austria': [0], 'Australia': [0], 'Belgium': [0], 'Brazil': [0], 'Switzerland': [0], 'Germany': [0, 1], 'Denmark': [0], 'England': [0, 1, 2, 3, 4], 'Spain': [0, 1], 'Finland': [0], 'France': [0, 1], 'Netherlands': [0], 'Italy': [0, 1], 'Japan': [0], 'Norway': [0], 'Poland': [0], 'Portugal': [0], 'Russia': [0], 'Scotland': [0], 'Sweden': [0], 'Turkey': [0]}

        #for continent in ['europa','amerika','asien']:
         #   search_soup = BeautifulSoup(requests.get('https://www.transfermarkt.com/wettbewerbe/{}'.format(continent), headers={'User-Agent': 'Mozilla/5.0'}).content, 'lxml').find_all('area')
         #   for country in search_soup:
          #      if country['title'] in countries_link_dict.keys():
           #         countries_link_dict[country['title']] = country['href']
        ###RESULT OF ABOVE###
        #countries_link_dict = {'Argentina': '/wettbewerbe/national/wettbewerbe/9', 'Austria': '/wettbewerbe/national/wettbewerbe/127', 'Australia': '/wettbewerbe/national/wettbewerbe/12', 'Belgium': '/wettbewerbe/national/wettbewerbe/19', 'Brazil': '/wettbewerbe/national/wettbewerbe/26', 'Switzerland': '/wettbewerbe/national/wettbewerbe/148', 'Germany': '/wettbewerbe/national/wettbewerbe/40', 'Denmark': '/wettbewerbe/national/wettbewerbe/39', 'England': '/wettbewerbe/national/wettbewerbe/189', 'Spain': '/wettbewerbe/national/wettbewerbe/157', 'Finland': '/wettbewerbe/national/wettbewerbe/49', 'France': '/wettbewerbe/national/wettbewerbe/50', 'Netherlands': '/wettbewerbe/national/wettbewerbe/122', 'Italy': '/wettbewerbe/national/wettbewerbe/75', 'Japan': '/wettbewerbe/national/wettbewerbe/77', 'Norway': '/wettbewerbe/national/wettbewerbe/125', 'Poland': '/wettbewerbe/national/wettbewerbe/135', 'Portugal': '/wettbewerbe/national/wettbewerbe/136', 'Russia': '/wettbewerbe/national/wettbewerbe/141', 'Scotland': '/wettbewerbe/national/wettbewerbe/190', 'Sweden': '/wettbewerbe/national/wettbewerbe/147', 'Turkey': '/wettbewerbe/national/wettbewerbe/174'}

        #league_name_code_list = []
        #for country, link in countries_link_dict.items():
         #   search_soup = BeautifulSoup(requests.get('https://www.transfermarkt.com{}'.format(link), headers={'User-Agent': 'Mozilla/5.0'}).content, 'lxml').find_all('tr',{'class':['odd','even']})
         #   for e in countries_dict[country]:
          #      league_name_code_list.append(list(search_soup[e].find_all('a')[1]['href'].split('/')[i] for i in [1,4]))
        ###RESULT OF ABOVE###
        league_name_code_list = [['superliga', 'AR1N'], ['bundesliga', 'A1'], ['a-league', 'AUS1'], ['jupiler-pro-league', 'BE1'], ['campeonato-brasileiro-serie-a', 'BRA1'], ['super-league', 'C1'], ['bundesliga', 'L1'], ['2-bundesliga', 'L2'], ['superligaen', 'DK1'], ['premier-league', 'GB1'], ['championship', 'GB2'], ['league-one', 'GB3'], ['league-two', 'GB4'], ['national-league', 'CNAT'], ['laliga', 'ES1'], ['laliga2', 'ES2'], ['veikkausliiga', 'FI1'], ['ligue-1', 'FR1'], ['ligue-2', 'FR2'], ['eredivisie', 'NL1'], ['serie-a', 'IT1'], ['serie-b', 'IT2'], ['j1-league', 'JAP1'], ['eliteserien', 'NO1'], ['pko-ekstraklasa', 'PL1'], ['liga-nos', 'PO1'], ['premier-liga', 'RU1'], ['scottish-premiership', 'SC1'], ['allsvenskan', 'SE1'], ['super-lig', 'TR1']]

        league_name_code_season = {}
        for i, league, name_code in zip(tqdm.tqdm(range(len(league_link_dict_std.keys()))), league_link_dict_std.keys(), league_name_code_list):
            search_soup = BeautifulSoup(requests.get('https://www.transfermarkt.com/{}/startseite/wettbewerb/{}'.format(name_code[0],name_code[1]), headers={'User-Agent': 'Mozilla/5.0'}).content, 'lxml').find('select',{'name':'saison_id'}).find_all({'option':'value'})
            for season in search_soup:
                search_soup = BeautifulSoup(requests.get('https://www.transfermarkt.com/{}/startseite/wettbewerb/{}/plus/?saison_id={}'.format(name_code[0],name_code[1],season['value']), headers={'User-Agent': 'Mozilla/5.0'}).content, 'lxml').find('div',{'class':'large-4'}).find_all('div',{'class':'box'})
                if len(search_soup) == 1:
                    pass
                if len(search_soup) > 1:
                    search_soup = BeautifulSoup(requests.get('https://www.transfermarkt.com/{}/startseite/wettbewerb/{}/plus/?saison_id={}'.format(name_code[0],name_code[1],season['value']), headers={'User-Agent': 'Mozilla/5.0'}).content, 'lxml').find_all('tbody')[-1].find_all('tr')[0]
                    if search_soup.find_all('td')[-3].text != '-':
                        league_name_code_season[league] = [name_code[0],name_code[1],season['value']]
                        time.sleep(0.1)
                        break
        self.dict_name_code_season = league_name_code_season
        
    def get_links_players(self):
        return self.dict_name_code_season

class links_downloaded():
    def __init__(self):
        self.dict_standings = {'Argentina': 'latest.asp?league=argentina', 'Austria': 'latest.asp?league=austria', 'Australia': 'latest.asp?league=australia', 'Belgium': 'latest.asp?league=belgium', 'Brazil': 'latest.asp?league=brazil', 'Switzerland': 'latest.asp?league=switzerland', 'Germany - Bundesliga': 'latest.asp?league=germany', 'Germany - 2. Bundesliga': 'latest.asp?league=germany2', 'Denmark': 'latest.asp?league=denmark', 'England - Premier League': 'latest.asp?league=england', 'England - Championship': 'latest.asp?league=england2', 'England - League One': 'latest.asp?league=england3', 'England - League Two': 'latest.asp?league=england4', 'England - National league': 'latest.asp?league=england5', 'Spain - La Liga': 'latest.asp?league=spain', 'Spain - La Liga 2': 'latest.asp?league=spain2', 'Finland': 'latest.asp?league=finland', 'France - Ligue 1': 'latest.asp?league=france', 'France - Ligue 2': 'latest.asp?league=france2', 'Netherlands': 'latest.asp?league=netherlands', 'Italy - Serie A': 'latest.asp?league=italy', 'Italy - Serie B': 'latest.asp?league=italy2', 'Japan': 'latest.asp?league=japan', 'Norway': 'latest.asp?league=norway', 'Poland': 'latest.asp?league=poland', 'Portugal': 'latest.asp?league=portugal', 'Russia': 'latest.asp?league=russia', 'Scotland': 'latest.asp?league=scotland', 'Sweden': 'latest.asp?league=sweden', 'Turkey': 'latest.asp?league=turkey'}
        self.dict_name_code_season = {'Argentina': ['superliga', 'AR1N', '2019'], 'Austria': ['bundesliga', 'A1', '2019'], 'Australia': ['a-league', 'AUS1', '2019'], 'Belgium': ['jupiler-pro-league', 'BE1', '2019'], 'Brazil': ['campeonato-brasileiro-serie-a', 'BRA1', '2018'], 'Switzerland': ['super-league', 'C1', '2019'], 'Germany - Bundesliga': ['bundesliga', 'L1', '2019'], 'Germany - 2. Bundesliga': ['2-bundesliga', 'L2', '2019'], 'Denmark': ['superligaen', 'DK1', '2019'], 'England - Premier League': ['premier-league', 'GB1', '2019'], 'England - Championship': ['championship', 'GB2', '2019'], 'England - League One': ['league-one', 'GB3', '2019'], 'England - League Two': ['league-two', 'GB4', '2019'], 'England - National league': ['national-league', 'CNAT', '2019'], 'Spain - La Liga': ['laliga', 'ES1', '2019'], 'Spain - La Liga 2': ['laliga2', 'ES2', '2019'], 'Finland': ['veikkausliiga', 'FI1', '2018'], 'France - Ligue 1': ['ligue-1', 'FR1', '2019'], 'France - Ligue 2': ['ligue-2', 'FR2', '2019'], 'Netherlands': ['eredivisie', 'NL1', '2019'], 'Italy - Serie A': ['serie-a', 'IT1', '2019'], 'Italy - Serie B': ['serie-b', 'IT2', '2019'], 'Japan': ['j1-league', 'JAP1', '2018'], 'Norway': ['eliteserien', 'NO1', '2018'], 'Poland': ['pko-ekstraklasa', 'PL1', '2019'], 'Portugal': ['liga-nos', 'PO1', '2019'], 'Russia': ['premier-liga', 'RU1', '2019'], 'Scotland': ['scottish-premiership', 'SC1', '2019'], 'Sweden': ['allsvenskan', 'SE1', '2018'], 'Turkey': ['super-lig', 'TR1', '2019']}
    
    def get_links_standings(self):
        return self.dict_standings
    def get_links_players(self):
        return self.dict_name_code_season

class league_table_dataframe():
    def __init__(self, league, league_link):
        #Get headers and build the dict
        headers = []
        search_soup = BeautifulSoup(requests.get('https://www.soccerstats.com/{}'.format(league_link), headers={'User-Agent': 'Mozilla/5.0'}).content, 'lxml').find_all('tr',{'class':'even'})[0].find_all('span')
        for header in search_soup:
            headers.append(header['title'])
        del headers[9]
        headers.insert(1,'Team')
        num_headers = ['Ranking', 'Games played', 'Games won', 'Draws', 'Games lost', 'Goals For', 'Goals Against', 'Goal Difference', 'Points']
        float_headers = ['Points Per Game', 'Points Per Game in the last 8 matches']
        percent_headers = ['% clean sheets (matches with no goal conceded)', '% failed to score (matches with no goal scored)']
        #######
        standings_table = {}
        for header in headers:
            standings_table[header] = []
        #Get rows
        search_soup = BeautifulSoup(requests.get('https://www.soccerstats.com/{}'.format(league_link), headers={'User-Agent': 'Mozilla/5.0'}).content, 'lxml').find('table',{'id':'btable'}).find_all('tr',{'class':'odd'})
        for element in search_soup:
            row = []
            row_line = element.text.split('\n')
            for e in row_line:
                e = e.split('\r')
                for e1 in e:
                    e1 = e1.split('\xa0')
                    if e1[0] != '':
                        row.append(e1[0])
                    if len(e1) == 2:
                        row.append(e1[1])
        #Merge rows with headers in the dict
            if len(row) == 13:
                for header in standings_table.keys():
                    if header == 'Ranking': 
                        standings_table['Ranking'].append('League not started')
                    if header == 'Team':
                        standings_table['Team'].append(row[0])
                    if header != 'Ranking' and header != 'Team':
                        standings_table[header].append(0)
            else:
                for header, row_e in zip(standings_table.keys(),(row[:10]+row[-4:])):
                    if header in num_headers:
                        standings_table[header].append(int(row_e))
                    if header in float_headers:
                        standings_table[header].append(float(row_e))
                    if header in percent_headers:
                        standings_table[header].append(int(row_e.strip('%'))/100)
                    if header not in num_headers and header not in percent_headers and header not in float_headers:
                        standings_table[header].append(row_e)
        #Generate the dataframe and change the types
        standings = pd.DataFrame(standings_table)
        standings.name = league
        self.standings = standings
    def get_standings(self):
        return self.standings

class scorer_assist_gkeepr_df():
    def __init__(self, league, league_name_code_season, scorer_assist_gkeep):
        if scorer_assist_gkeep == 'Scorers':
            comp_links = ['torschuetzenliste', '', 'altersklasse/alle/detailpos//plus/1']
        if scorer_assist_gkeep == 'Assistants':
            comp_links = ['torschuetzenliste', 'ajax/yw1', 'altersklasse/alle/detailpos//plus/1/sort/vorlagen.desc']
        if scorer_assist_gkeep == 'Goalkeepers':
            comp_links = ['weisseWeste', '', 'plus/1']
        #Extract the headers
        headers_as_list = []
        search_soup = BeautifulSoup(requests.get('https://www.transfermarkt.com/{}/{}/wettbewerb/{}/{}/saison_id/{}/{}'.format(league_name_code_season[0], comp_links[0], league_name_code_season[1], comp_links[1], league_name_code_season[2], comp_links[2]), headers={'User-Agent': 'Mozilla/5.0'}).content, 'lxml').find('div',{'class':'responsive-table'}).find('thead').find_all('th')    
        for header in search_soup:
            if header.text == '\xa0':
                headers_as_list.append(header.find('span')['title'])
            if header.text == '':
                headers_as_list.append(header.find('div')['title'])
            else:
                if header.text != '\xa0':
                    headers_as_list.append(header.text)
        if scorer_assist_gkeep == 'Scorers' or scorer_assist_gkeep == 'Assistants':
            headers_as_list.insert(2,'Position')
        if scorer_assist_gkeep == 'Goalkeepers':
            splitted_headers = headers_as_list[1].split('/')
            headers_as_list[1] = splitted_headers[0]
            headers_as_list.insert(2,splitted_headers[1])
        #Extract the rows 
        search_soup = BeautifulSoup(requests.get('https://www.transfermarkt.com/{}/{}/wettbewerb/{}/{}/saison_id/{}/{}'.format(league_name_code_season[0], comp_links[0], league_name_code_season[1], comp_links[1], league_name_code_season[2], comp_links[2]), headers={'User-Agent': 'Mozilla/5.0'}).content, 'lxml').find('div',{'class':'responsive-table'}).find('tbody').find_all('tr',{'class':['odd','even']})
        dataframe_dict = {}
        for row in search_soup:
            row_as_list = []
            for element in row.find_all('td'):
                if '\n' in element.text or element.text == '':
                    #GET NAME
                    if element.find('a',{'class':'spielprofil_tooltip'}) != None:
                        if element.find('a',{'class':'spielprofil_tooltip'})['title'] not in row_as_list:
                            row_as_list.append(element.find('a',{'class':'spielprofil_tooltip'})['title'])
                    #GET NATION
                    if element.find('img',{'class':'flaggenrahmen'}) != None:
                        if element.find('img',{'class':'flaggenrahmen'})['title'] not in row_as_list:
                            row_as_list.append(element.find('img',{'class':'flaggenrahmen'})['title'])
                    #GET CLUB
                    if element.find('img',{'class':''}) != None:
                        if element.find('img',{'class':''})['alt'] not in row_as_list:
                            row_as_list.append(element.find('img',{'class':''})['alt'])  
                else:
                    if element.text == '-':
                        row_as_list.append(0)
                    else:
                        row_as_list.append(element.text)
            if len(row_as_list) > 1:
                for header, column in zip(headers_as_list,row_as_list):
                    if header not in dataframe_dict.keys():
                        dataframe_dict[header] = [column]
                    else:
                        dataframe_dict[header].append(column)
        #Change formats of numerical columns
        def formater(header):
            formatted_list = []
            for data in dataframe_dict[header]:
                if header == 'Age (today)':
                    formatted_list.append(data.split("(")[0].strip())
                if header == 'Minutes played':
                    formatted_list.append(data.split("'")[0].replace('.',''))
                if header == 'Minutes per goal':
                    formatted_list.append(data.split("'")[0])
                if header == 'Goals per match':
                    formatted_list.append(data.replace(',','.'))
                if header == 'minutes of play':
                    formatted_list.append(data.replace('.',''))
                if header == 'Percentage':
                    formatted_list.append(float(data.split('%')[0].strip().replace(',','.'))/100)        
            return formatted_list
        columns_to_format = ['Age (today)', 'Minutes played', 'Minutes per goal', 'Goals per match', 'minutes of play', 'Percentage']
        for column in columns_to_format:
            if column in dataframe_dict.keys():
                dataframe_dict[column] = formater(column)
        #Generate the dataframe
        scorer_assist_gkeep_df = pd.DataFrame(dataframe_dict)
        if scorer_assist_gkeep == 'Scorers' or scorer_assist_gkeep == 'Assistants': 
            scorer_assist_gkeep_df[['#','Age (today)','Appearances','Assists','Penalty kicks','Minutes played','Minutes per goal','Goals per match','Goals']] = scorer_assist_gkeep_df[['#','Age (today)','Appearances','Assists','Penalty kicks','Minutes played','Minutes per goal','Goals per match','Goals']].apply(pd.to_numeric)
            scorer_assist_gkeep_df.name = league
            self.scorer_assist_gkeep_df = scorer_assist_gkeep_df
        if scorer_assist_gkeep == 'Goalkeepers':
            scorer_assist_gkeep_df[['#','Matches','Clean Sheets','Goals conceded','minutes of play','Minutes per goal against','Percentage']] = scorer_assist_gkeep_df[['#','Matches','Clean Sheets','Goals conceded','minutes of play','Minutes per goal against','Percentage']].apply(pd.to_numeric)
            scorer_assist_gkeep_df.name = league
            self.scorer_assist_gkeep_df = scorer_assist_gkeep_df

    def get_ranking(self):
        return self.scorer_assist_gkeep_df

class all_dataframes_together():
    def __init__(self, league_links_dict, Stand_Scor_Asst_Gkeep):
        list_dataframes = []
        for i, league, links in zip(tqdm.tqdm(range(len(league_links_dict))), league_links_dict.keys(), list(league_links_dict.values())):
            if Stand_Scor_Asst_Gkeep == 'Standings':
                dataframe = league_table_dataframe(league, links).get_standings()
            if Stand_Scor_Asst_Gkeep == 'Scorers':
                dataframe = scorer_assist_gkeepr_df(league, links, 'Scorers').get_ranking()
            if Stand_Scor_Asst_Gkeep == 'Assistants':
                dataframe = scorer_assist_gkeepr_df(league, links, 'Assistants').get_ranking()
            if Stand_Scor_Asst_Gkeep == 'Goalkeepers':
                dataframe = scorer_assist_gkeepr_df(league, links, 'Goalkeepers').get_ranking()
            dataframe.insert(0, 'League', list(itertools.repeat(dataframe.name, len(dataframe))), True)
            list_dataframes.append(dataframe)
            time.sleep(0.1)  
        new_dataframe = pd.concat(list_dataframes)
        new_dataframe.name = 'All {} together'.format(Stand_Scor_Asst_Gkeep)
        self.new_dataframe = new_dataframe
    
    def mod_dataframe(self):
        return self.new_dataframe

class dataframe_to_excel():
    def __init__(self, dataframe, Standings_Scorers_Assistants_Goalkeepers):
        sheets_ordered = ['Argentina', 'Austria', 'Australia', 'Belgium', 'Brazil', 'Switzerland', 'Germany - Bundesliga', 'Germany - 2. Bundesliga', 'Denmark', 'England - Premier League', 'England - Championship', 'England - League One', 'England - League Two', 'England - National league', 'Spain - La Liga', 'Spain - La Liga 2', 'Finland', 'France - Ligue 1', 'France - Ligue 2', 'Netherlands', 'Italy - Serie A', 'Italy - Serie B', 'Japan', 'Norway', 'Poland', 'Portugal', 'Russia', 'Scotland', 'Sweden', 'Turkey', 'All {} together'.format(Standings_Scorers_Assistants_Goalkeepers)]
        out_path = '{}/{}'.format(os.getcwd(),'Saved dataframes')
        sheet_name = dataframe.name
        writer = pd.ExcelWriter('{}/{}.xlsx'.format(out_path, Standings_Scorers_Assistants_Goalkeepers), engine='openpyxl')
        wb = load_workbook('{}/{}.xlsx'.format(out_path, Standings_Scorers_Assistants_Goalkeepers)) 
        writer.book = wb
        if sheet_name in wb.sheetnames:
            excel_sheet = pd.read_excel('{}/{}.xlsx'.format(out_path, Standings_Scorers_Assistants_Goalkeepers), sheet_name=sheet_name)
            excel_sheet = excel_sheet.loc[:, excel_sheet.columns != 'Unnamed: 0']
            if dataframe.equals(excel_sheet) == True:
                self.new_file = None
            else:
                del wb[sheet_name]
                dataframe.to_excel(writer, sheet_name=sheet_name, startrow=0, startcol=0)
                wb._sheets = wb._sheets[:sheets_ordered.index(sheet_name)] + [wb._sheets[-1]] + wb._sheets[sheets_ordered.index(sheet_name):]
                del wb._sheets[-1]
                writer.save()
                writer.close()
                self.new_file = None
        else:
            dataframe.to_excel(writer, sheet_name='{}'.format(sheet_name))
            if 'Sheet1' in wb.sheetnames: 
                del wb['Sheet1']
            writer.save()
            writer.close()
            self.new_file = None                 
    
    def save_dataframe(self):
        self.new_file

